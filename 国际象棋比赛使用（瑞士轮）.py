import random
import math
from collections import defaultdict
import os
from datetime import datetime
import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False
try:
    import matplotlib.pyplot as plt
    from matplotlib.table import Table
    from matplotlib.backends.backend_pdf import PdfPages

    HAS_MPL = True
except Exception:
    HAS_MPL = False
import json  # 新增：用于保存/读取状态

class Player:
    def __init__(self, name, team="", gender="", category=""):
        self.name = name
        self.team = team
        self.gender = gender  # 新增：性别
        self.category = category  # 新增：组别/类别
        self.number = None  # 报名编号（按录入顺序，从1开始）
        self.score = 0.0
        self.opponents = []
        self.sonneborn_berger = 0.0
        self.wins = 0
        self.losses = 0
        self.draws = 0
        self.bye = False
        self.rank = 0
        self.color_history = []  # 记录先后手历史
        self.matches = []  # 每轮对局记录：{round, opponent, color, result}
        self.absent_streak = 0  # 连续缺席轮数
        self.disqualified = False  # 是否被取消资格
        self.dq_round = None  # 取消资格生效轮次（用于标注）
    
    def add_result(self, opponent, result, color=None, round_no=None, note: str = None):
        """记录单局结果。
        opponent: 对手姓名或 "BYE"
        result: 从本方视角 1/0.5/0
        color: 'W'/'B'/None
        round_no: 轮次编号（1 开始）；None 时按已有记录推断
        """
        self.opponents.append(opponent)
        if color:
            self.color_history.append(color)

        if round_no is None:
            round_no = len(self.matches) + 1
        self.matches.append({
            'round': round_no,
            'opponent': opponent,
            'color': color,
            'result': result,
            'note': note or '',
        })

        if result == 1:
            self.score += 1
            self.wins += 1
        elif result == 0.5:
            self.score += 0.5
            self.draws += 1
        elif result == 0:
            self.losses += 1
    
    def calculate_sonneborn_berger(self, players_dict):
        """按标准定义计算 SB 小分：胜加对手总分，和加对手总分一半，负不加；BYE 不计入。"""
        self.sonneborn_berger = 0.0
        for m in self.matches:
            opp = m['opponent']
            if opp == "BYE":
                continue
            if opp not in players_dict:
                continue
            opp_score = players_dict[opp].score
            res = m['result']
            if res == 1:
                self.sonneborn_berger += opp_score
            elif res == 0.5:
                self.sonneborn_berger += 0.5 * opp_score
    
    def get_last_color(self):
        """获取最后一次使用的颜色"""
        return self.color_history[-1] if self.color_history else None
    
    def get_color_balance(self):
        """计算颜色平衡（正数表示白棋多，负数表示黑棋多）"""
        white_count = self.color_history.count('W')
        black_count = self.color_history.count('B')
        return white_count - black_count
    
    def __str__(self):
        return f"{self.name} ({self.team}) - 积分: {self.score}, 胜: {self.wins}, 和: {self.draws}, 负: {self.losses}"

    # 重新根据 matches 重算本人的基础统计（用于撤销与修改当轮结果）
    def recompute_from_matches(self):
        self.score = 0.0
        self.wins = 0
        self.draws = 0
        self.losses = 0
        # 重建先后手与对手列表
        self.color_history = []
        self.opponents = []
        # 按轮次顺序恢复，以保证缺席计数等的顺序正确
        for m in sorted(self.matches, key=lambda m: m['round']):
            self.opponents.append(m['opponent'])
            if m.get('color'):
                self.color_history.append(m['color'])
            res = m.get('result')
            if res == 1:
                self.score += 1
                self.wins += 1
            elif res == 0.5:
                self.score += 0.5
                self.draws += 1
            elif res == 0:
                self.losses += 1
        # 重新计算连续缺席轮数（以“缺席”备注为准；BYE 与“对手缺席”均不计缺席）
        streak = 0
        for m in sorted(self.matches, key=lambda m: m['round']):
            if m.get('note') == '缺席':
                streak += 1
            else:
                streak = 0
        self.absent_streak = streak

class SwissTournament:
    def __init__(self):
        self.players = []
        self.rounds = 0
        self.current_round = 0
        self.pairings_history = []
        self.team_players = defaultdict(list)
        self.team_scores = defaultdict(float)
        self.use_teams = False
        self.tournament_name = ""
        self.tournament_location = ""   # 新增：地点
        self.tournament_category = ""   # 新增：组别/类别（用于输出目录/文件名后缀）
        self._name_counts = defaultdict(int)  # 用于姓名去重
        # 取消资格相关
        self.defer_disqualify = True  # 取消资格是否在本轮结束后生效
        self._pending_disqualifications = set()  # 待在本轮结束后处理的选手名
        # 团体设置
        self.team_top_n = None  # 团体计分取最佳 N 人（可设置或赛后输入）
        # BYE 自动计分设置
        self.auto_bye_scoring = True
        # 导出根目录
        self.out_dir = None
        self.pending_pairings = None  # 新增：保存“本轮对阵表已生成但尚未录入结果”的对阵，以便断点续录
    
    def clear_screen(self):
        """清屏"""
        os.system('cls' if os.name == 'nt' else 'clear')
    
    def input_tournament_settings(self):
        """输入比赛设置"""
        self.clear_screen()
        print("=== 国际象棋瑞士轮比赛设置 ===")
        self.tournament_name = input("请输入比赛名称: ").strip()
        
        # 新增：地点与组别（可选）
        loc = input("请输入比赛地点（可选，回车跳过）: ").strip()
        cat = input("请输入比赛组别/类别（可选，回车跳过）: ").strip()
        self.tournament_location = loc
        self.tournament_category = cat
        
        # 选择是否启用团体模式
        team_choice = input("是否启用团体模式? (y/n): ").lower()
        self.use_teams = (team_choice == 'y')
        if self.use_teams:
            # 团体最佳 N 人设置（可留空赛后再输入）
            topn = input("团体计分采用最佳 N 人，请输入 N（回车跳过，赛后再输入，默认3）: ").strip()
            if topn:
                try:
                    nval = int(topn)
                    if nval > 0:
                        self.team_top_n = nval
                except ValueError:
                    pass
        
        # 输入比赛轮数
        while True:
            try:
                self.rounds = int(input("请输入比赛轮数: "))
                if self.rounds > 0:
                    break
                else:
                    print("轮数必须大于0")
            except ValueError:
                print("请输入有效的数字")
        
        # 取消资格生效时机
        dq_choice = input("选手连缺两轮是否在本轮结束后再生效取消资格? (y/n，默认 y): ").lower().strip() or 'y'
        self.defer_disqualify = (dq_choice == 'y')
        
        # BYE 自动计分设置
        bye_choice = input("是否自动记分处理轮空（BYE）? (y/n，默认 y): ").lower().strip() or 'y'
        self.auto_bye_scoring = (bye_choice == 'y')

    def input_players(self):
        """输入参赛选手信息"""
        self.clear_screen()
        print("=== 输入参赛选手 ===")
        use_excel = (input("是否从 Excel 导入选手编号与姓名? (y/n, 输入 t 生成样板, 默认 n): ").strip().lower() or 'n')
        if use_excel == 't':
            # 生成 Excel 样板
            path = input("请输入要保存样板的路径（含文件名 .xlsx，回车为当前目录 '参赛选手样板.xlsx'）: ").strip().strip('"')
            if not path:
                path = os.path.join(os.getcwd(), "参赛选手样板.xlsx")
            try:
                self.export_excel_template(path)
                print(f"[提示] 已生成样板: {path}")
            except Exception as e:
                print(f"[错误] 生成样板失败：{e}")
            # 改动：生成样板后不直接 return，而是询问是否立即用该样板导入或继续手动输入
            choice = input("是否现在用此样板导入选手? (y 导入 / 任意键 返回手动录入): ").strip().lower() or ''
            if choice == 'y':
                try:
                    self.import_players_from_excel(path)
                    return
                except Exception as e:
                    print(f"[提示] 用样板导入失败：{e}，将进入手动录入。")
            # 如果不导入或导入失败，继续后续的手动录入流程（不直接 return）
        if use_excel == 'y':
            path = input("请输入 Excel 文件路径（.xlsx）: ").strip().strip('"')
            try:
                self.import_players_from_excel(path)
                return
            except Exception as e:
                print(f"[提示] Excel 导入失败：{e}，转为手动录入。")
        n = int(input("请输入参赛选手人数: "))
        
        for i in range(n):
            print(f"\n输入第{i+1}位选手信息:")
            name = input("姓名: ")
            # 姓名去重校验/自动改名
            name = self.ensure_unique_name(name)
            
            if self.use_teams:
                team = input("队伍: ")
            else:
                team = ""  # 个人赛模式，队伍为空
            gender = input("性别（可选）: ").strip()
            category = input("组别/类别（可选）: ").strip()
            
            player = Player(name, team, gender, category)
            player.number = i + 1
            self.players.append(player)
            
            if self.use_teams and team:
                self.team_players[team].append(player)

    def export_excel_template(self, path: str):
        """生成一个 Excel 导入样板，包含常用列：编号/姓名/队伍/性别/组别"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except Exception:
            raise RuntimeError("需要安装 openpyxl 以生成 Excel 样板（pip install openpyxl）")
        wb = Workbook()
        ws = wb.active
        headers = ["编号", "姓名", "队伍", "性别", "组别"]
        ws.append(headers)
        # 添加示例行
        ws.append([1, "张三", "队伍A", "男", "公开组"])
        ws.append([2, "李四", "队伍B", "女", "公开组"])
        wb.save(path)

    def import_players_from_excel(self, path: str):
        """从 Excel 导入选手：支持列名 ‘编号’、‘姓名’、‘队伍’、'性别'、'组别'(可选)。编号用于 player.number。"""
        import os
        if not os.path.exists(path):
            raise FileNotFoundError("文件不存在")
        try:
            import openpyxl
        except Exception:
            raise RuntimeError("需要安装 openpyxl 库以读取 .xlsx 文件")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # 寻找表头行
        headers = {}
        header_row = None
        for r in range(1, min(ws.max_row, 20) + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            texts = [str(v).strip() if v is not None else '' for v in vals]
            if any(x in ("编号", "姓名", "队伍", "性别", "组别") for x in texts):
                header_row = r
                for c, t in enumerate(texts, start=1):
                    if t:
                        headers[t] = c
                break
        if header_row is None or "姓名" not in headers:
            raise ValueError("未找到包含‘姓名’的表头")
        col_no = headers.get("编号")
        col_name = headers.get("姓名")
        col_team = headers.get("队伍")
        col_gender = headers.get("性别")
        col_category = headers.get("组别")
        # 清空并重新构建
        self.players = []
        self.team_players = defaultdict(list)
        self._name_counts = defaultdict(int)
        count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            name_cell = ws.cell(row=r, column=col_name).value
            if name_cell is None or str(name_cell).strip() == '':
                continue
            name = self.ensure_unique_name(str(name_cell).strip())
            team = ''
            gender = ''
            category = ''
            if col_team:
                tv = ws.cell(row=r, column=col_team).value
                team = (str(tv).strip() if tv is not None else '')
            if col_gender:
                gv = ws.cell(row=r, column=col_gender).value
                gender = (str(gv).strip() if gv is not None else '')
            if col_category:
                cv = ws.cell(row=r, column=col_category).value
                category = (str(cv).strip() if cv is not None else '')
            p = Player(name, team, gender, category)
            if col_no:
                nv = ws.cell(row=r, column=col_no).value
                try:
                    p.number = int(str(nv).strip()) if nv is not None and str(nv).strip() != '' else None
                except Exception:
                    p.number = None
            if p.number is None:
                p.number = len(self.players) + 1
            self.players.append(p)
            if self.use_teams and team:
                self.team_players[team].append(p)
            count += 1
        # 编号唯一性：若重复则按出现顺序重排
        nums = [p.number for p in self.players]
        if len(nums) != len(set(nums)):
            for i, p in enumerate(self.players, start=1):
                p.number = i
        print(f"已从 Excel 导入有效选手 {len(self.players)} 人（读取 {count} 行）。")

    def ensure_unique_name(self, name: str) -> str:
        """确保姓名唯一，若重复则自动添加后缀 (2), (3), ..."""
        base = name.strip()
        if not base:
            base = "选手"
        existing = {p.name for p in self.players}
        if base not in existing and self._name_counts[base] == 0:
            self._name_counts[base] = 1
            return base
        # 已存在，生成下一个可用序号
        count = max(self._name_counts[base], 1)
        while True:
            count += 1
            candidate = f"{base} ({count})"
            if candidate not in existing:
                self._name_counts[base] = count
                print(f"[提示] 姓名重复，已自动更名为：{candidate}")
                return candidate
    
    def get_ranking_criteria(self, player):
        """获取排名标准元组"""
        players_dict = {p.name: p for p in self.players}
        opponent_scores = 0
        for opp_name in player.opponents:
            if opp_name != "BYE" and opp_name in players_dict:
                opponent_scores += players_dict[opp_name].score
        
        return (
            player.score,
            player.sonneborn_berger,
            self.get_head_to_head(player),
            opponent_scores,
            player.wins
        )
    
    def get_head_to_head(self, player):
        """计算直胜关系（简化版）"""
        # 在实际比赛中需要记录具体的对战结果来计算
        return 0
    
    def rank_players(self):
        """对选手进行排名"""
        players_dict = {player.name: player for player in self.players}
        for player in self.players:
            player.calculate_sonneborn_berger(players_dict)
        
        sorted_players = sorted(self.players, key=self.get_ranking_criteria, reverse=True)
        
        for i, player in enumerate(sorted_players):
            player.rank = i + 1
        
        return sorted_players
    
    def can_pair(self, player1, player2, proposed_pairings):
        """检查两个选手是否可以配对"""
        # 如果启用团体模式，同队不能配对
        if self.use_teams and player1.team == player2.team:
            return False
        
        # 已经对战过不能配对
        if player2.name in player1.opponents:
            return False
        
        # 检查是否已经在当前轮次的配对中
        for pair in proposed_pairings:
            if player1.name in pair or player2.name in pair:
                return False
        
        return True
    
    def assign_colors(self, pairings):
        """分配颜色，尽量保持颜色平衡"""
        colored_pairings = []
        
        for player1_name, player2_name in pairings:
            if player2_name == "BYE":
                colored_pairings.append((player1_name, "BYE", None))
                continue
            
            player1 = next(p for p in self.players if p.name == player1_name)
            player2 = next(p for p in self.players if p.name == player2_name)
            
            # 计算颜色平衡
            balance1 = player1.get_color_balance()
            balance2 = player2.get_color_balance()
            
            if balance1 < balance2:
                # player1白棋，player2黑棋
                colored_pairings.append((player1_name, player2_name, 'W'))
            elif balance1 > balance2:
                # player1黑棋，player2白棋
                colored_pairings.append((player1_name, player2_name, 'B'))
            else:
                # 平衡相同，随机分配
                color = random.choice(['W', 'B'])
                colored_pairings.append((player1_name, player2_name, color))
        
        return colored_pairings
    
    def swiss_pairing(self):
        """瑞士制配对算法"""
        if self.current_round == 0:
            # 第一轮随机配对
            pairings = self.random_pairing()
        else:
            # 按积分分组
            sorted_players = self.rank_players()
            # 仅从未取消资格的选手中进行配对
            available_players = [p for p in sorted_players if not p.disqualified]

            # 新增：若人数为奇数，先“预选”本轮 BYE（优先从未轮空者中选择），并从可配对列表中剔除
            preselected_bye = None
            if len(available_players) % 2 == 1:
                # _choose_bye_player 会优先选未轮空者；若全都轮空过，则可能返回已轮空者
                preselected_bye = self._choose_bye_player(available_players)
                if preselected_bye is None:
                    preselected_bye = random.choice(available_players)
                # 从可配对池中移除该选手
                available_players = [p for p in available_players if p.name != preselected_bye.name]

            score_groups = defaultdict(list)
            for player in available_players:
                score_groups[player.score].append(player)
            
            proposed_pairings = []
            paired_players = set()
            
            # 从高分到低分进行配对
            for score in sorted(score_groups.keys(), reverse=True):
                group = score_groups[score][:]
                
                while group:
                    player1 = group.pop(0)
                    if player1.name in paired_players:
                        continue
                    
                    # 寻找合适的对手
                    opponent_found = False
                    for i, player2 in enumerate(group):
                        if self.can_pair(player1, player2, proposed_pairings):
                            proposed_pairings.append((player1.name, player2.name))
                            paired_players.add(player1.name)
                            paired_players.add(player2.name)
                            group.pop(i)
                            opponent_found = True
                            break
                    
                    # 如果当前分数组找不到对手，尝试从低分组找：优先靠近的低分组（分数高的低分组）
                    if not opponent_found:
                        for lower_score in sorted((s for s in score_groups.keys() if s < score), reverse=True):
                            for player2 in score_groups[lower_score]:
                                if (player2.name not in paired_players and 
                                    self.can_pair(player1, player2, proposed_pairings)):
                                    proposed_pairings.append((player1.name, player2.name))
                                    paired_players.add(player1.name)
                                    paired_players.add(player2.name)
                                    score_groups[lower_score].remove(player2)
                                    opponent_found = True
                                    break
                            if opponent_found:
                                break
            
            # 安全网：完成未配对者的兜底配对（已确保总人数为偶数，不再新增 BYE）
            unpaired_players = [p for p in available_players if p.name not in paired_players]

            def try_pair_with_constraints(players_list, allow_same_team=False, allow_repeat=False):
                """尽量在给定约束下完成两两配对；保证有限步退出。"""
                remaining = players_list[:]
                made_pairs = []
                i = 0
                while i < len(remaining) - 1:
                    a = remaining[i]
                    match_idx = -1
                    for j in range(i + 1, len(remaining)):
                        b = remaining[j]
                        if not allow_same_team and self.use_teams and a.team == b.team:
                            continue
                        if not allow_repeat and (b.name in a.opponents or a.name in b.opponents):
                            continue
                        match_idx = j
                        break
                    if match_idx != -1:
                        b = remaining.pop(match_idx)
                        a = remaining.pop(i)
                        made_pairs.append((a.name, b.name))
                    else:
                        i += 1
                return made_pairs, remaining

            staging = [
                (False, False),
                (True, False),
                (False, True),
                (True, True),
            ]
            remaining = unpaired_players[:]
            for allow_same_team, allow_repeat in staging:
                if len(remaining) < 2:
                    break
                pairs, remaining = try_pair_with_constraints(remaining, allow_same_team, allow_repeat)
                for a, b in pairs:
                    proposed_pairings.append((a, b))
                    paired_players.add(a)
                    paired_players.add(b)

            if len(remaining) >= 2:
                for k in range(0, len(remaining) - 1, 2):
                    proposed_pairings.append((remaining[k].name, remaining[k+1].name))
                    paired_players.add(remaining[k].name)
                    paired_players.add(remaining[k+1].name)

            # 新增：将预选 BYE 追加到结果末尾，严格避免同一选手二次 BYE（除非所有人都已轮空）
            if preselected_bye:
                if preselected_bye.bye:
                    # 若不可避免（所有人都已轮空过），给出提示
                    if any(not p.bye for p in [p for p in self.players if not p.disqualified]):
                        print(f"[提示] 已尽量避免二次轮空，但出现极端情况。")
                proposed_pairings.append((preselected_bye.name, "BYE"))
                preselected_bye.bye = True

            pairings = proposed_pairings
        
        # 分配颜色
        return self.assign_colors(pairings)
    
    def random_pairing(self):
        """第一轮随机配对"""
        # 仅从未取消资格的选手中进行随机配对（理论上第一轮不会有DQ）
        shuffled_players = [p for p in self.players if not p.disqualified]
        random.shuffle(shuffled_players)
        
        pairings = []
        for i in range(0, len(shuffled_players), 2):
            if i + 1 < len(shuffled_players):
                pairings.append((shuffled_players[i].name, shuffled_players[i+1].name))
            else:
                # 改动：确保即使 _choose_bye_player 返回 None 也能安全处理
                p = self._choose_bye_player([shuffled_players[i]])
                if p is None:
                    p = shuffled_players[i]
                pairings.append((p.name, "BYE"))
                p.bye = True
        
        # 覆盖性自检：确保每位选手出现一次
        names = {p.name for p in self.players if not p.disqualified}
        covered = set()
        for a, b in pairings:
            covered.add(a)
            if b != "BYE":
                covered.add(b)
        if names != covered:
            print("[警告] 第一轮配对覆盖性异常，已尝试修正。")
            # 简单兜底：将遗漏者两两配对，若剩 1 人则给 BYE
            missing = [n for n in names if n not in covered]
            while len(missing) >= 2:
                a = missing.pop()
                b = missing.pop()
                pairings.append((a, b))
            if len(missing) == 1:
                pairings.append((missing[0], "BYE"))
        return pairings
    
    def input_single_result(self, pairing_index, colored_pairings):
        """输入单个台次的比赛结果"""
        self.clear_screen()
        print(f"=== 输入第{self.current_round + 1}轮比赛结果 ===")
        
        player1_name, player2_name, color = colored_pairings[pairing_index]
        players_dict = {player.name: player for player in self.players}
        
        print(f"\n台次 {pairing_index + 1}:")
        if color == 'W':
            white_player = player1_name
            black_player = player2_name
        elif color == 'B':
            white_player = player2_name
            black_player = player1_name
        else:  # 轮空
            p = players_dict[player1_name]
            pscore = p.score if player1_name in players_dict else 0
            num = f"#{p.number} " if getattr(p, 'number', None) else ""
            print(f"{num}{player1_name} [轮空, 分 {pscore:.1f}]")
            
        if color in ('W','B'):
            wp = players_dict[white_player]
            bp = players_dict[black_player]
            white_score = wp.score if white_player in players_dict else 0
            black_score = bp.score if black_player in players_dict else 0
            wnum = f"#{wp.number} " if getattr(wp, 'number', None) else ""
            bnum = f"#{bp.number} " if getattr(bp, 'number', None) else ""
            print(f"白方: {wnum}{white_player} (分 {white_score:.1f}) vs 黑方: {bnum}{black_player} (分 {black_score:.1f})")
        
        if player2_name == "BYE":
            print("轮空自动计为胜利")
            players_dict[player1_name].add_result("BYE", 1, None, round_no=self.current_round + 1, note='BYE')
            # BYE 不是缺席，重置该选手缺席计数
            players_dict[player1_name].absent_streak = 0
            return True
        
        while True:
            try:
                # 提示区：再次回显当前积分，便于快速核对
                if color in ('W','B'):
                    print(f"\n当前积分 | 白: {white_player} {white_score:.1f} 分  | 黑: {black_player} {black_score:.1f} 分")
                else:
                    print(f"\n当前积分 | 轮空选手: {player1_name} {pscore:.1f} 分")
                print("请输入结果:")
                print("1: 白方胜")
                print("2: 黑方胜") 
                print("3: 和棋")
                print("4: 白方缺席（黑方判胜）")
                print("5: 黑方缺席（白方判胜）")
                choice = input("选择 (1/2/3/4/5): ")
                
                if choice == "1":
                    if color == 'W':
                        players_dict[player1_name].add_result(player2_name, 1, 'W', round_no=self.current_round + 1)
                        players_dict[player2_name].add_result(player1_name, 0, 'B', round_no=self.current_round + 1)
                    else:
                        players_dict[player2_name].add_result(player1_name, 1, 'W', round_no=self.current_round + 1)
                        players_dict[player1_name].add_result(player2_name, 0, 'B', round_no=self.current_round + 1)
                    # 正常对局，双方缺席计数清零
                    players_dict[player1_name].absent_streak = 0
                    players_dict[player2_name].absent_streak = 0
                    break
                elif choice == "2":
                    if color == 'W':
                        players_dict[player1_name].add_result(player2_name, 0, 'W', round_no=self.current_round + 1)
                        players_dict[player2_name].add_result(player1_name, 1, 'B', round_no=self.current_round + 1)
                    else:
                        players_dict[player2_name].add_result(player1_name, 0, 'W', round_no=self.current_round + 1)
                        players_dict[player1_name].add_result(player2_name, 1, 'B', round_no=self.current_round + 1)
                    # 正常对局，双方缺席计数清零
                    players_dict[player1_name].absent_streak = 0
                    players_dict[player2_name].absent_streak = 0
                    break
                elif choice == "3":
                    players_dict[player1_name].add_result(player2_name, 0.5, color, round_no=self.current_round + 1)
                    players_dict[player2_name].add_result(player1_name, 0.5, 'B' if color == 'W' else 'W', round_no=self.current_round + 1)
                    # 和棋两人均非缺席，重置缺席计数
                    players_dict[player1_name].absent_streak = 0
                    players_dict[player2_name].absent_streak = 0
                    break
                elif choice == "4":
                    # 白方缺席，黑方判胜
                    white = player1_name if color == 'W' else player2_name
                    black = player2_name if color == 'W' else player1_name
                    players_dict[white].add_result(black, 0, 'W', round_no=self.current_round + 1, note='缺席')
                    players_dict[black].add_result(white, 1, 'B', round_no=self.current_round + 1, note='对手缺席')
                    players_dict[white].absent_streak += 1
                    players_dict[black].absent_streak = 0
                    self._check_and_disqualify(players_dict[white])
                    break
                elif choice == "5":
                    # 黑方缺席，白方判胜
                    white = player1_name if color == 'W' else player2_name
                    black = player2_name if color == 'W' else player1_name
                    players_dict[black].add_result(white, 0, 'B', round_no=self.current_round + 1, note='缺席')
                    players_dict[white].add_result(black, 1, 'W', round_no=self.current_round + 1, note='对手缺席')
                    players_dict[black].absent_streak += 1
                    players_dict[white].absent_streak = 0
                    self._check_and_disqualify(players_dict[black])
                    break
                else:
                    print("无效选择，请重新输入")
            except Exception as e:
                print(f"输入错误: {e}")
        
        print("结果输入成功!")
        input("\n按回车键继续...")
        return True

    def _check_and_disqualify(self, player: Player):
        """连续两轮缺席则取消参赛资格。"""
        if player.absent_streak >= 2 and not player.disqualified:
            if self.defer_disqualify:
                if player.name not in self._pending_disqualifications:
                    self._pending_disqualifications.add(player.name)
                    print(f"[通知] 选手 {player.name} 连续两轮缺席，将在本轮结束后取消参赛资格。")
            else:
                print(f"[通知] 选手 {player.name} 连续两轮缺席，已取消参赛资格！")
                self.disqualify_player(player.name)

    def disqualify_player(self, name: str):
        """标记取消资格（保留在选手列表与统计中，但不再参与后续配对）。"""
        for p in self.players:
            if p.name == name:
                p.disqualified = True
                if p.dq_round is None:
                    p.dq_round = self.current_round  # 在当前轮结束时或即时生效时记录
                break
    
    def display_current_round_pairings(self, colored_pairings, results_entered):
        """显示当前轮对阵表和已输入结果的情况"""
        self.clear_screen()
        print(f"=== 第{self.current_round + 1}轮对阵表 ===")
        print("● 已输入结果  ○ 未输入结果\n")
        
        players_dict = {player.name: player for player in self.players}
        
        for i, (player1_name, player2_name, color) in enumerate(colored_pairings):
            status = "●" if results_entered[i] else "○"
            
            if player2_name == "BYE":
                pscore = players_dict[player1_name].score if player1_name in players_dict else 0
                p = players_dict[player1_name]
                num = f"#{p.number} " if getattr(p, 'number', None) else ""
                print(f"{status} 台次 {i+1}: {num}{player1_name} [轮空, 分 {pscore:.1f}]")
            else:
                if color == 'W':
                    white_player = player1_name
                    black_player = player2_name
                else:
                    white_player = player2_name
                    black_player = player1_name
                wp = players_dict[white_player]
                bp = players_dict[black_player]
                white_score = wp.score if white_player in players_dict else 0
                black_score = bp.score if black_player in players_dict else 0
                wnum = f"#{wp.number} " if getattr(wp, 'number', None) else ""
                bnum = f"#{bp.number} " if getattr(bp, 'number', None) else ""
                
                # 显示结果（如果已输入）
                result_str = ""
                if results_entered[i]:
                    p1 = players_dict[player1_name]
                    p2 = players_dict[player2_name]
                    # 用 matches 精确查找最近一次对战结果
                    def last_match_result(p, opp):
                        for m in reversed(p.matches):
                            if m['opponent'] == opp:
                                return m
                        return None
                    if white_player == player1_name:
                        m = last_match_result(p1, player2_name)
                        if m:
                            if m['result'] == 1 and m['color'] == 'W':
                                result_str = " 1-0"
                            elif m['result'] == 0 and m['color'] == 'W':
                                result_str = " 0-1"
                            else:
                                result_str = " ½-½"
                    else:
                        m = last_match_result(p2, player1_name)
                        if m:
                            if m['result'] == 1 and m['color'] == 'W':
                                result_str = " 1-0"
                            elif m['result'] == 0 and m['color'] == 'W':
                                result_str = " 0-1"
                            else:
                                result_str = " ½-½"
                
                print(f"{status} 台次 {i+1}: {wnum}{white_player} (白, 分 {white_score:.1f}) vs {bnum}{black_player} (黑, 分 {black_score:.1f}){result_str}")
    
    def manage_round_results(self, colored_pairings):
        """管理一轮比赛结果的输入"""
        results_entered = [False] * len(colored_pairings)
        # 自动处理本轮所有 BYE（可配置）
        if self.auto_bye_scoring:
            players_dict = {player.name: player for player in self.players}
            for i, (p1, p2, color) in enumerate(colored_pairings):
                if p2 == "BYE" and not results_entered[i]:
                    players_dict[p1].add_result("BYE", 1, None, round_no=self.current_round + 1, note='BYE')
                    players_dict[p1].absent_streak = 0
                    results_entered[i] = True
        
        while not all(results_entered):
            self.display_current_round_pairings(colored_pairings, results_entered)
            
            print(f"\n已输入 {sum(results_entered)}/{len(results_entered)} 个台次的结果")
            print("\n选项:")
            print("1. 输入指定台次结果")
            print("2. 查看当前排名")
            if self.use_teams:
                print("3. 查看团体排名")
                print("4. 修改已输入台次结果")
            else:
                print("3. 修改已输入台次结果")
            print("0. 完成本轮所有结果输入")
            
            choice = input("\n请选择: ")
            
            if choice == "1":
                try:
                    pairing_index = int(input("请输入台次号: ")) - 1
                    if 0 <= pairing_index < len(colored_pairings):
                        if not results_entered[pairing_index]:
                            self.input_single_result(pairing_index, colored_pairings)
                            results_entered[pairing_index] = True
                        else:
                            print("该台次结果已输入!")
                            input("按回车键继续...")
                    else:
                        print("无效的台次号!")
                        input("按回车键继续...")
                except ValueError:
                    print("请输入有效的数字!")
                    input("按回车键继续...")
            
            elif choice == "2":
                self.display_ranking()
                input("\n按回车键继续...")
            
            elif choice == "3" and self.use_teams:
                self.display_team_ranking()
                input("\n按回车键继续...")

            elif (choice == "3" and not self.use_teams) or (choice == "4" and self.use_teams):
                # 修改已输入台次结果
                try:
                    pairing_index = int(input("请输入要修改的台次号: ")) - 1
                    if 0 <= pairing_index < len(colored_pairings):
                        if results_entered[pairing_index]:
                            self._undo_pairing_result(pairing_index, colored_pairings, results_entered)
                            print("已撤销该台次结果，现在可以重新录入。")
                            input("按回车键继续...")
                        else:
                            print("该台次尚未输入结果，无需修改！")
                            input("按回车键继续...")
                    else:
                        print("无效的台次号！")
                        input("按回车键继续...")
                except ValueError:
                    print("请输入有效的数字！")
                    input("按回车键继续...")
            
            elif choice == "0":
                if all(results_entered):
                    break
                else:
                    confirm = input("还有未输入的结果，确定要结束本轮吗? (y/n): ")
                    if confirm.lower() == 'y':
                        break
        
        # 保存本轮记录
        self.pairings_history.append((self.current_round + 1, colored_pairings, results_entered))
        self.current_round += 1

    def _remove_match_for_player(self, player: Player, opp_name: str, round_no: int):
        # 仅移除当轮该对手的记录（若存在）
        idx_to_remove = None
        for idx, m in enumerate(player.matches):
            if m.get('round') == round_no and m.get('opponent') == opp_name:
                idx_to_remove = idx
                break
        if idx_to_remove is not None:
            player.matches.pop(idx_to_remove)
            player.recompute_from_matches()

    def _refresh_disqualification_after_edit(self):
        # 撤销/修改后，刷新待取消资格队列；并在“即时生效”模式下必要时回滚本轮错误触发的 DQ
        if self.defer_disqualify:
            self._pending_disqualifications = set()
            for p in self.players:
                if (not p.disqualified) and p.absent_streak >= 2:
                    self._pending_disqualifications.add(p.name)
        else:
            for p in self.players:
                # 若本轮曾即时标记DQ，但修改后已不满足连续两轮缺席，则回滚
                if p.disqualified and p.dq_round == self.current_round and p.absent_streak < 2:
                    p.disqualified = False
                    p.dq_round = None

    def _undo_pairing_result(self, pairing_index, colored_pairings, results_entered):
        # 撤销指定台次的已输入结果（限当前轮），并刷新缺席/DQ状态
        players_dict = {player.name: player for player in self.players}
        p1, p2, color = colored_pairings[pairing_index]
        round_no = self.current_round + 1
        if p2 == 'BYE':
            if p1 in players_dict:
                self._remove_match_for_player(players_dict[p1], 'BYE', round_no)
        else:
            if p1 in players_dict:
                self._remove_match_for_player(players_dict[p1], p2, round_no)
            if p2 in players_dict:
                self._remove_match_for_player(players_dict[p2], p1, round_no)
        # 标记该台次为未输入
        results_entered[pairing_index] = False
        # 刷新取消资格挂起/回滚
        self._refresh_disqualification_after_edit()
    
    def display_ranking(self):
        """显示个人排名"""
        self.clear_screen()
        print(f"=== 第{self.current_round}轮后个人排名 ===")
        ranked_players = self.rank_players()
        
        print(f"{'排名':<4} {'姓名':<18} {'队伍':<10} {'积分':<6} {'索尼伯恩':<10} {'胜':<4} {'和':<4} {'负':<4} {'先后手':<8}")
        print("-" * 80)
        
        for i, player in enumerate(ranked_players):
            color_balance = player.get_color_balance()
            color_str = f"W{color_balance}" if color_balance >= 0 else f"B{-color_balance}"
            dq_tag = (f" (DQ@{player.dq_round})" if player.disqualified else "")
            name_disp = (player.name + dq_tag)
            print(f"{i+1:<4} {name_disp:<18} {player.team:<10} {player.score:<6} "
                  f"{player.sonneborn_berger:<10.2f} {player.wins:<4} {player.draws:<4} {player.losses:<4} {color_str:<8}")
    
    def display_team_ranking(self):
        """显示团体排名"""
        self.clear_screen()
        print(f"=== 第{self.current_round}轮后团体排名 ===")
        
        if not self.use_teams:
            print("本次比赛未启用团体模式")
            return
        
        while True:
            try:
                n = int(input("请输入计算团体成绩的人数 n: "))
                if n > 0:
                    break
                else:
                    print("人数必须大于0")
            except ValueError:
                print("请输入有效的数字")
        
        team_scores = defaultdict(list)
        
        # 收集每个队伍选手的排名
        for player in self.players:
            team_scores[player.team].append(player.rank)
        
        # 计算每个队伍的最佳n人排名和
        team_results = []
        for team, ranks in team_scores.items():
            if len(ranks) >= n:
                best_ranks = sorted(ranks)[:n]
                rank_sum = sum(best_ranks)
                team_score = sum(player.score for player in self.team_players[team])
                team_results.append((team, rank_sum, team_score, len(ranks)))
        
        # 按排名和排序，相同则比较队伍总积分
        team_results.sort(key=lambda x: (x[1], -x[2]))
        
        print(f"\n{'排名':<4} {'队伍':<15} {'最佳{n}人排名和':<15} {'队伍总积分':<10} {'人数':<4}")
        print("-" * 60)
        
        for i, (team, rank_sum, team_score, count) in enumerate(team_results):
            print(f"{i+1:<4} {team:<15} {rank_sum:<15} {team_score:<10.1f} {count:<4}")
    
    def run_tournament(self):
        """运行整个比赛"""
        self.clear_screen()
        print("=== 国际象棋瑞士轮比赛管理系统 ===")
        # 新增：开局支持从状态文件恢复
        try:
            resume = (input("是否从保存的状态文件恢复比赛? (y/n, 默认 n): ").strip().lower() or 'n') == 'y'
        except Exception:
            resume = False
        if resume:
            path = input("请输入状态文件路径（.json）: ").strip().strip('"')
            if os.path.exists(path) and self.load_state_from_file(path):
                print("[提示] 已从状态文件恢复。")
                input("按回车键继续...")
            else:
                print("[提示] 恢复失败，进入常规新比赛流程。")
                input("按回车键继续...")
                resume = False

        # 比赛设置与选手录入（仅在未恢复时进行）
        if not resume:
            # 比赛设置
            self.input_tournament_settings()
            # 输入选手
            self.input_players()

        # 开始比赛
        while self.current_round < self.rounds:
            self.clear_screen()
            print(f"=== {self.tournament_name} ===")
            print(f"当前轮次: {self.current_round + 1}/{self.rounds}")

            # 生成对阵表（若存在 pending_pairings 则直接使用）
            print("\n正在生成对阵表...")
            out_dir = self.ensure_out_dir()
            if self.pending_pairings:
                colored_pairings = self.pending_pairings
            else:
                colored_pairings = self.swiss_pairing()
            # 导出本轮对阵表图片
            pairings_img = self.export_round_pairings_image(self.current_round + 1, colored_pairings, out_dir)
            # 新增：保存“对阵后、录入前”的状态（包含 pending_pairings）
            self.pending_pairings = colored_pairings
            self.save_state(out_dir, stage="pre")  # R{round+1:02d}_对阵后_状态.json

            # 输入比赛结果
            self.manage_round_results(colored_pairings)
            # 本轮录入结束，pending_pairings 清空
            self.pending_pairings = None

            # 显示本轮后排名
            self.display_ranking()
            # 若选择“本轮结束后”再取消资格，则此处应用待处理的取消资格（在显示完本轮排名之后）
            self.apply_pending_disqualifications()

            # 导出本轮结果图 + 个人/团体排名图，并生成当轮 PDF
            results_img = self.export_round_results_image(self.current_round, colored_pairings, out_dir)
            # 导出当前总排名图（覆盖同名或用当轮命名）
            rank_filename = f"R{self.current_round:02d}_个人总排名.png"
            # 改动：传入 round_no，在图片标题标注“第X轮后个人总排名”
            self.export_overall_ranking_image(out_dir, filename=rank_filename, round_no=self.current_round)
            team_img = None
            if self.use_teams:
                team_filename = f"R{self.current_round:02d}_团体总排名.png"
                # 确保 N 可用
                if self.team_players:
                    max_allowed = max(1, min(len(v) for v in self.team_players.values()))
                else:
                    max_allowed = 1
                n = self.team_top_n if (self.team_top_n and self.team_top_n <= max_allowed) else min(3, max_allowed)
                self.export_team_ranking_image(out_dir, top_n=n, filename=team_filename)
                # 修正：团队排名图保存时带有比赛名称前缀
                team_img = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_{team_filename}")
            # 生成当轮 PDF
            # 修正：总排名图保存时带有比赛名称前缀
            rank_img_path = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_{rank_filename}")
            imgs = [p for p in [pairings_img, results_img, rank_img_path, team_img] if p]
            # 修正：当轮 PDF 文件名包含比赛名称
            pdf_path = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_R{self.current_round:02d}_当轮结果与排名.pdf")
            self.export_pdf_from_images(imgs, pdf_path)

            # 新增：保存“本轮排名后”的状态（含已完成的当轮）
            self.save_state(out_dir, stage="post")  # R{round:02d}_排名后_状态.json

            input("\n按回车键查看下一轮...")

            # 是否继续下一轮
            if self.current_round < self.rounds:
                cont = input("\n是否继续下一轮? (y/n): ")
                if cont.lower() != 'y':
                    break

        # 显示最终排名
        self.clear_screen()
        print("=== 比赛结束 ===")
        print(f"比赛名称: {self.tournament_name}")
        print(f"总轮次: {self.rounds}")
        
        self.display_ranking()
        # 比赛完全结束：如果仍有待处理取消资格，一般无需应用（仅影响下一轮配对），可清空队列
        self._pending_disqualifications.clear()
        
        if self.use_teams:
            print("\n" + "="*50)
            self.display_team_ranking()
        
        print("\n感谢使用国际象棋瑞士轮比赛管理系统!")

        # 导出图片与汇总 PDF（统一到同一输出目录）
        out_dir = self.ensure_out_dir()
        self.export_player_images(out_dir)
        # 总排名表（赛后总排名，不带轮次）
        self.export_overall_ranking_image(out_dir)
        # 团体排名表（若启用团体）
        if self.use_teams:
            # 若未在赛前设置，则赛后询问一次
            if self.team_top_n is None:
                try:
                    topn_post = input("赛后导出：请输入团体计分的最佳 N 人（回车默认 3）: ").strip()
                    if topn_post:
                        nval = int(topn_post)
                        if nval > 0:
                            self.team_top_n = nval
                except Exception:
                    pass
            # 计算一个安全的 N（不超过任何队伍人数）
            if self.team_players:
                max_allowed = max(1, min(len(v) for v in self.team_players.values()))
            else:
                max_allowed = 1
            n = self.team_top_n if (self.team_top_n and self.team_top_n <= max_allowed) else min(3, max_allowed)
            self.export_team_ranking_image(out_dir, top_n=n)
        # 汇总 PDF（把该目录下所有生成的 PNG 合并为一个总 PDF）
        self.export_pdf_summary(out_dir)
        # 新增：导出“总排名与对阵情况”的专用总PDF（顺序更清晰）
        self.export_final_summary_pdf(out_dir)

    def save_state(self, out_dir: str, stage: str):
        """保存当前比赛状态为 JSON。
        stage: 'pre'（对阵后、录入前）或 'post'（当轮排名后）
        """
        data = {
            'meta': {
                'tournament_name': self.tournament_name,
                'tournament_location': getattr(self, 'tournament_location', ''),
                'tournament_category': getattr(self, 'tournament_category', ''),
                'rounds': self.rounds,
                'current_round': self.current_round,
                'use_teams': self.use_teams,
                'team_top_n': self.team_top_n,
                'defer_disqualify': self.defer_disqualify,
                'auto_bye_scoring': self.auto_bye_scoring,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'stage': stage,
                'out_dir': out_dir,
            },
            'players': [],
            'pairings_history': [],
            'pending_disqualifications': list(self._pending_disqualifications),
            'pending_pairings': self.pending_pairings if self.pending_pairings else None,
        }
        # 玩家
        for p in self.players:
            data['players'].append({
                'name': p.name,
                'team': p.team,
                'gender': p.gender,
                'category': p.category,
                'number': p.number,
                'matches': p.matches,  # 已是 JSON 友好的字典列表
                'disqualified': p.disqualified,
                'dq_round': p.dq_round,
                'bye': p.bye,
            })
        # 对阵历史（转为字典）
        for rno, cps, entered in self.pairings_history:
            data['pairings_history'].append({
                'round': rno,
                'pairings': cps,          # [ [p1, p2, color], ... ]
                'entered': list(entered), # [bool,...]
            })
        # 文件名
        safe = self._safe_filename(self.tournament_name)
        if stage == 'pre':
            r = self.current_round + 1
            fname = f"{safe}_R{r:02d}_对阵后_状态.json"
        else:
            r = self.current_round
            fname = f"{safe}_R{r:02d}_排名后_状态.json"
        path = os.path.join(out_dir, fname)
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"[已保存状态] {path}")
        except Exception as e:
            print(f"[警告] 保存状态失败：{e}")

    def load_state_from_file(self, path: str) -> bool:
        """从 JSON 状态文件恢复。成功返回 True。"""
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"[错误] 读取状态文件失败：{e}")
            return False
        try:
            meta = data.get('meta', {})
            # 基本元数据
            self.tournament_name = meta.get('tournament_name', '')
            self.tournament_location = meta.get('tournament_location', '')
            self.tournament_category = meta.get('tournament_category', '')
            self.rounds = int(meta.get('rounds', 0))
            self.current_round = int(meta.get('current_round', 0))
            self.use_teams = bool(meta.get('use_teams', False))
            self.team_top_n = meta.get('team_top_n', None)
            self.defer_disqualify = bool(meta.get('defer_disqualify', True))
            self.auto_bye_scoring = bool(meta.get('auto_bye_scoring', True))
            self.out_dir = meta.get('out_dir') or self.ensure_out_dir()
            os.makedirs(self.out_dir, exist_ok=True)

            # 玩家
            self.players = []
            self.team_players = defaultdict(list)
            raw_players = data.get('players', [])
            for pd in raw_players:
                p = Player(
                    name=pd.get('name', ''),
                    team=pd.get('team', ''),
                    gender=pd.get('gender', ''),
                    category=pd.get('category', '')
                )
                p.number = pd.get('number', None)
                # 恢复对局
                p.matches = list(pd.get('matches', []))
                # 由 matches 反推基础统计（积分/先后手/缺席串等）
                p.recompute_from_matches()
                # 取消资格/轮空标记
                p.disqualified = bool(pd.get('disqualified', False))
                p.dq_round = pd.get('dq_round', None)
                p.bye = bool(pd.get('bye', False))
                self.players.append(p)
                if self.use_teams and p.team:
                    self.team_players[p.team].append(p)

            # 对阵历史
            self.pairings_history = []
            for item in data.get('pairings_history', []):
                rno = item.get('round')
                cps = item.get('pairings', [])
                entered = item.get('entered', [])
                self.pairings_history.append((rno, cps, entered))

            # 待处理取消资格
            self._pending_disqualifications = set(data.get('pending_disqualifications', []))
            # 待录入对阵（断点续录）
            self.pending_pairings = data.get('pending_pairings', None)

            # 重新计算 SB/排名，确保一致
            self.rank_players()
            return True
        except Exception as e:
            print(f"[错误] 恢复状态失败：{e}")
            return False

    def _safe_filename(self, s: str) -> str:
        invalid = '<>:"/\\|?*\n\r\t'
        for ch in invalid:
            s = s.replace(ch, '_')
        return s.strip() or 'player'

    def export_player_images(self, out_dir: str = None):
        """导出每位选手的比赛与小分情况为图片（PNG）。"""
        if not HAS_MPL:
            print("[提示] 未检测到 matplotlib，无法导出图片。可通过 pip 安装 matplotlib 后重试。")
            return None
        if out_dir is None:
            # 改动：默认使用统一的输出目录，避免新建另一个目录而导致汇总 PDF 漏图
            out_dir = self.ensure_out_dir()
        os.makedirs(out_dir, exist_ok=True)

        # 按最终排名排序，确保对手分数为最终分
        ranked = self.rank_players()
        players_dict = {p.name: p for p in self.players}

        for idx, p in enumerate(ranked, start=1):
            fig = plt.figure(figsize=(10, 7), dpi=150)
            # 上：累计积分曲线
            ax_curve = fig.add_axes([0.08, 0.70, 0.88, 0.22])
            ms = sorted(p.matches, key=lambda m: m['round']) if p.matches else []
            rounds = [m['round'] for m in ms]
            scores = []
            cum = 0.0
            for m in ms:
                if m['opponent'] == 'BYE':
                    cum += 1.0
                else:
                    cum += m['result']
                scores.append(cum)
            ax_curve.plot(rounds, scores, marker='o')
            ax_curve.set_xlim(left=0.5, right=(rounds[-1] if rounds else 1)+0.5)
            ax_curve.set_ylim(bottom=0)
            ax_curve.grid(True, linestyle='--', alpha=0.4)
            ax_curve.set_ylabel('累计积分')
            dq_tag = f" (DQ@{p.dq_round})" if p.disqualified else ""
            # 标题中加入比赛名称
            title = (f"{self.tournament_name} | {p.name}{dq_tag}"
                     f"{' [' + p.team + ']' if p.team else ''}  | 最终排名 #{p.rank}  | 积分 {p.score:.1f}"
                     f"  | SB小分 {p.sonneborn_berger:.2f}  | 战绩 {p.wins}-{p.draws}-{p.losses}")
            ax_curve.set_title(title, fontweight='bold', fontsize=12, loc='left')

            # 下：对局与 SB 表格
            ax = fig.add_axes([0.02, 0.06, 0.96, 0.60])
            ax.axis('off')

            headers = ["轮次", "先后手", "对手", "对手排名", "结果", "对手总分", "此局SB贡献"]
            rows = []
            for m in ms:
                opp = m['opponent']
                color = m['color'] or ''
                res = m['result']
                note = m.get('note', '')
                if opp == 'BYE':
                    opp_score = ''
                    opp_rank = ''
                    sb = ''
                    res_str = '1 (BYE)'
                else:
                    opp_obj = players_dict.get(opp)
                    opp_score = opp_obj.score if opp_obj else ''
                    opp_rank = opp_obj.rank if opp_obj else ''
                    sb = (opp_score if res == 1 else (0.5 * opp_score if res == 0.5 else 0)) if opp_score != '' else ''
                    if res == 1:
                        res_str = '胜'
                    elif res == 0.5:
                        res_str = '和'
                    else:
                        res_str = '负'
                    if note:
                        res_str += f" ({note})"
                color_str = {'W': '白', 'B': '黑'}.get(color, '')
                rows.append([m['round'], color_str, opp, opp_rank, res_str, opp_score, f"{sb:.2f}" if sb != '' else ''])

            color_balance = p.get_color_balance()
            cb_str = f"颜色平衡：{'白多' if color_balance>0 else ('黑多' if color_balance<0 else '均衡')} (W{p.color_history.count('W')}/B{p.color_history.count('B')})"
            bye_rounds = [m['round'] for m in ms if m['opponent']=='BYE']
            br_str = f"轮空：{('无' if not bye_rounds else '第' + ','.join(map(str, bye_rounds)) + '轮')}"
            ax.text(0.02, 0.95, cb_str + '    ' + br_str, transform=ax.transAxes, fontsize=10)

            tab = Table(ax, bbox=[0.02, 0.05, 0.96, 0.82])
            # 适当压缩列宽以容纳“对手排名”列
            col_widths = [0.07, 0.09, 0.24, 0.10, 0.16, 0.14, 0.14]
            for j, h in enumerate(headers):
                tab.add_cell(0, j, width=col_widths[j], height=0.06, text=h, loc='center', facecolor='#ECEFF1')
            for i, row in enumerate(rows, start=1):
                for j, val in enumerate(row):
                    tab.add_cell(i, j, width=col_widths[j], height=0.06, text=str(val), loc='center')
            ax.add_table(tab)

            footer = datetime.now().strftime('生成时间：%Y-%m-%d %H:%M:%S')
            ax.text(0.02, 0.01, footer, transform=ax.transAxes, fontsize=8, color='#666666')

            fname = f"{idx:02d}_{self._safe_filename(self.tournament_name)}_{self._safe_filename(p.name)}.png"
            save_path = os.path.join(out_dir, fname)
            plt.savefig(save_path, bbox_inches='tight')
            plt.close(fig)

        print(f"\n已导出每位选手的比赛与小分图片到：{out_dir}")
        return out_dir

    def export_overall_ranking_image(self, out_dir: str, filename: str = "00_个人总排名.png", round_no: int = None):
        if not HAS_MPL:
            return
        ranked = self.rank_players()
        fig = plt.figure(figsize=(11, 8), dpi=150)
        ax = fig.add_subplot(111)
        ax.axis('off')
        # 改动：标题随 round_no 变化
        if round_no is not None and round_no > 0:
            ax.set_title(f'{self.tournament_name} - 第{round_no}轮后个人总排名', fontweight='bold', fontsize=14)
        else:
            ax.set_title(f'{self.tournament_name} - 个人总排名', fontweight='bold', fontsize=14)
        headers = ["排名", "姓名", "队伍", "积分", "SB小分", "胜", "和", "负", "颜色平衡"]
        rows = []
        for p in ranked:
            cb = p.get_color_balance()
            cb_str = f"W{p.color_history.count('W')}/B{p.color_history.count('B')} ({'白多' if cb>0 else ('黑多' if cb<0 else '均衡')})"
            name_disp = p.name + (f" (DQ@{p.dq_round})" if p.disqualified else "")
            rows.append([p.rank, name_disp, p.team, f"{p.score:.1f}", f"{p.sonneborn_berger:.2f}", p.wins, p.draws, p.losses, cb_str])
        tab = Table(ax, bbox=[0.02, 0.05, 0.96, 0.90])
        col_widths = [0.07, 0.20, 0.20, 0.10, 0.12, 0.07, 0.07, 0.07, 0.10]
        for j, h in enumerate(headers):
            tab.add_cell(0, j, width=col_widths[j], height=0.05, text=h, loc='center', facecolor='#ECEFF1')
        for i, row in enumerate(rows, start=1):
            for j, val in enumerate(row):
                tab.add_cell(i, j, width=col_widths[j], height=0.05, text=str(val), loc='center')
        ax.add_table(tab)
        path = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_{filename}")
        plt.savefig(path, bbox_inches='tight')
        plt.close(fig)

    def export_team_ranking_image(self, out_dir: str, top_n: int = 3, filename: str = "01_团体总排名.png"):
        if not HAS_MPL or not self.use_teams:
            return
        # 先确保 rank 已更新
        ranked = self.rank_players()
        # 收集团队的选手排名与总分
        team_ranks = defaultdict(list)
        team_scores = defaultdict(float)
        for p in self.players:
            team_ranks[p.team].append(p.rank)
            team_scores[p.team] += p.score
        team_results = []
        for team, ranks in team_ranks.items():
            if not team:
                continue
            if len(ranks) >= top_n:
                best = sorted(ranks)[:top_n]
                team_results.append((team, sum(best), team_scores[team], len(ranks), best))
        # 排序（排名和升序，总积分降序）
        team_results.sort(key=lambda x: (x[1], -x[2]))
        fig = plt.figure(figsize=(11, 6), dpi=150)
        ax = fig.add_subplot(111)
        ax.axis('off')
        # 标题中加入比赛名称和取最佳人数
        ax.set_title(f'{self.tournament_name} - 团体总排名（取最佳{top_n}人）', fontweight='bold', fontsize=14)
        headers = ["排名", "队伍", f"最佳{top_n}人排名和", "队伍总积分", "参赛人数", f"最佳{top_n}名次"]
        rows = []
        for i, (team, rank_sum, tscore, count, best) in enumerate(team_results, start=1):
            rows.append([i, team, rank_sum, f"{tscore:.1f}", count, ','.join(map(str, best))])
        tab = Table(ax, bbox=[0.02, 0.10, 0.96, 0.80])
        col_widths = [0.07, 0.30, 0.18, 0.15, 0.12, 0.18]
        for j, h in enumerate(headers):
            tab.add_cell(0, j, width=col_widths[j], height=0.06, text=h, loc='center', facecolor='#ECEFF1')
        for i, row in enumerate(rows, start=1):
            for j, val in enumerate(row):
                tab.add_cell(i, j, width=col_widths[j], height=0.06, text=str(val), loc='center')
        ax.add_table(tab)
        path = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_{filename}")
        plt.savefig(path, bbox_inches='tight')
        plt.close(fig)

    def export_round_pairings_image(self, round_no: int, colored_pairings, out_dir: str = None):
        if not HAS_MPL:
            return None
        out_dir = out_dir or self.ensure_out_dir()
        fig = plt.figure(figsize=(10, 7), dpi=150)
        ax = fig.add_subplot(111)
        ax.axis('off')
        # 标题包含比赛名称
        ax.set_title(f"{self.tournament_name} - 第{round_no}轮 对阵表", fontweight='bold', fontsize=14)
        headers = ["台次", "白方", "黑方"]
        rows = []
        players_dict = {p.name: p for p in self.players}
        for i, (p1, p2, color) in enumerate(colored_pairings, start=1):
            if p2 == 'BYE':
                p = players_dict.get(p1)
                w = f"#{p.number} {p1}" if p and getattr(p, 'number', None) else p1
                rows.append([i, w, "BYE"])
            else:
                if color == 'W':
                    wp, bp = p1, p2
                else:
                    wp, bp = p2, p1
                w = f"#{players_dict[wp].number} {wp}" if wp in players_dict and getattr(players_dict[wp], 'number', None) else wp
                b = f"#{players_dict[bp].number} {bp}" if bp in players_dict and getattr(players_dict[bp], 'number', None) else bp
                rows.append([i, w, b])
        tab = Table(ax, bbox=[0.02, 0.08, 0.96, 0.86])
        col_widths = [0.10, 0.45, 0.45]
        for j, h in enumerate(headers):
            tab.add_cell(0, j, width=col_widths[j], height=0.06, text=h, loc='center', facecolor='#ECEFF1')
        for i, row in enumerate(rows, start=1):
            for j, val in enumerate(row):
                tab.add_cell(i, j, width=col_widths[j], height=0.06, text=str(val), loc='center')
        ax.add_table(tab)
        path = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_R{round_no:02d}_对阵表.png")
        plt.savefig(path, bbox_inches='tight')
        plt.close(fig)
        return path

    def export_round_results_image(self, round_no: int, colored_pairings, out_dir: str = None):
        if not HAS_MPL:
            return None
        out_dir = out_dir or self.ensure_out_dir()
        fig = plt.figure(figsize=(10, 7), dpi=150)
        ax = fig.add_subplot(111)
        ax.axis('off')
        # 标题包含比赛名称
        ax.set_title(f"{self.tournament_name} - 第{round_no}轮 对局结果", fontweight='bold', fontsize=14)
        headers = ["台次", "白方", "黑方", "结果"]
        rows = []
        players_dict = {p.name: p for p in self.players}
        for i, (p1, p2, color) in enumerate(colored_pairings, start=1):
            if p2 == 'BYE':
                p = players_dict.get(p1)
                w = f"#{p.number} {p1}" if p and getattr(p, 'number', None) else p1
                rows.append([i, w, "BYE", "1-0 (BYE)"])
            else:
                if color == 'W':
                    wp, bp = p1, p2
                else:
                    wp, bp = p2, p1
                w = f"#{players_dict[wp].number} {wp}" if wp in players_dict and getattr(players_dict[wp], 'number', None) else wp
                b = f"#{players_dict[bp].number} {bp}" if bp in players_dict and getattr(players_dict[bp], 'number', None) else bp
                # 查结果
                res_str = ""
                wp_obj = players_dict.get(wp)
                if wp_obj:
                    # 找到本轮与对手的记录
                    m = None
                    for mm in wp_obj.matches:
                        if mm.get('round') == round_no and mm.get('opponent') == bp:
                            m = mm
                            break
                    if m:
                        if m.get('result') == 1 and m.get('color') == 'W':
                            res_str = '1-0'
                        elif m.get('result') == 0 and m.get('color') == 'W':
                            res_str = '0-1'
                        else:
                            res_str = '½-½'
                    else:
                        res_str = '—'
                else:
                    res_str = '—'
                rows.append([i, w, b, res_str])
        tab = Table(ax, bbox=[0.02, 0.08, 0.96, 0.86])
        col_widths = [0.10, 0.40, 0.40, 0.10]
        for j, h in enumerate(headers):
            tab.add_cell(0, j, width=col_widths[j], height=0.06, text=h, loc='center', facecolor='#ECEFF1')
        for i, row in enumerate(rows, start=1):
            for j, val in enumerate(row):
                tab.add_cell(i, j, width=col_widths[j], height=0.06, text=str(val), loc='center')
        ax.add_table(tab)
        path = os.path.join(out_dir, f"{self._safe_filename(self.tournament_name)}_R{round_no:02d}_结果.png")
        plt.savefig(path, bbox_inches='tight')
        plt.close(fig)
        return path

    def export_pdf_from_images(self, image_paths, pdf_path: str):
        """将多张图片合并为一个 PDF。自动跳过不存在的图片."""
        if not HAS_MPL:
            return
        if not image_paths:
            return
        with PdfPages(pdf_path) as pdf:
            for path in image_paths:
                if not os.path.exists(path):
                    continue
                img = plt.imread(path)
                h, w = img.shape[:2]
                fig = plt.figure(figsize=(w/100, h/100), dpi=100)
                ax = fig.add_subplot(111)
                ax.imshow(img)
                ax.axis('off')
                pdf.savefig(fig, bbox_inches='tight')
                plt.close(fig)

    def export_pdf_summary(self, out_dir: str, pdf_name: str = None):
        """按固定顺序合并为瑞士轮汇总 PDF：
           1) 最终总排名 -> 2) 所有选手个人战绩图 -> 3) 每轮：对阵表 -> 结果 -> 当轮结束后排名"""
        if not HAS_MPL:
            print("[提示] 未检测到 matplotlib，无法导出 PDF 汇总。")
            return
        safe = self._safe_filename(self.tournament_name)
        if pdf_name is None:
            pdf_name = f"{safe}_瑞士轮汇总.pdf"
        pdf_path = os.path.join(out_dir, pdf_name)

        pages = []

        # 1) 最终总排名
        final_overall = os.path.join(out_dir, f"{safe}_00_个人总排名.png")
        if os.path.exists(final_overall):
            pages.append(final_overall)

        # 2) 所有选手个人战绩图（按前缀编号排序）
        try:
            file_list = [f for f in os.listdir(out_dir) if f.lower().endswith('.png')]
            def is_player_img(fn: str) -> bool:
                return len(fn) > 4 and fn[:2].isdigit() and fn[2] == '_' and fn[3:].startswith(f"{safe}_")
            player_imgs = sorted([f for f in file_list if is_player_img(f)])
            pages.extend([os.path.join(out_dir, f) for f in player_imgs])
        except Exception:
            pass

        # 3) 每轮：对阵表 -> 结果 -> 当轮结束后排名（Rxx_个人总排名.png）
        for r in range(1, self.current_round + 1):
            pair_img = os.path.join(out_dir, f"{safe}_R{r:02d}_对阵表.png")
            res_img  = os.path.join(out_dir, f"{safe}_R{r:02d}_结果.png")
            rank_img = os.path.join(out_dir, f"{safe}_R{r:02d}_个人总排名.png")
            if os.path.exists(pair_img): pages.append(pair_img)
            if os.path.exists(res_img):  pages.append(res_img)
            if os.path.exists(rank_img): pages.append(rank_img)

        if not pages:
            print("[提示] 未找到可汇总的图片。")
            return

        with PdfPages(pdf_path) as pdf:
            for path in pages:
                try:
                    img = plt.imread(path)
                except Exception:
                    continue
                h, w = img.shape[:2]
                fig = plt.figure(figsize=(w/100, h/100), dpi=100)
                ax = fig.add_subplot(111)
                ax.imshow(img)
                ax.axis('off')
                pdf.savefig(fig, bbox_inches='tight')
                plt.close(fig)
        print(f"已生成 PDF 汇总：{pdf_path}")

    def export_final_summary_pdf(self, out_dir: str):
        """生成包含‘最终个人总排名 + （可选）最终团体总排名 + 各轮对阵表/结果表 + 每位选手个人战绩图(置于最后)’的总PDF。"""
        if not HAS_MPL:
            print("[提示] 未检测到 matplotlib，无法导出最终总 PDF。")
            return
        safe = self._safe_filename(self.tournament_name)
        pages = []
        # 最终个人总排名（赛后导出的 00_个人总排名.png）
        final_overall = os.path.join(out_dir, f"{safe}_00_个人总排名.png")
        if os.path.exists(final_overall):
            pages.append(final_overall)
        # 最终团体总排名（可选）
        final_team = os.path.join(out_dir, f"{safe}_01_团体总排名.png")
        if self.use_teams and os.path.exists(final_team):
            pages.append(final_team)
        # 各轮对阵表与结果表
        for r in range(1, self.current_round + 1):
            pairings = os.path.join(out_dir, f"{safe}_R{r:02d}_对阵表.png")
            results = os.path.join(out_dir, f"{safe}_R{r:02d}_结果.png")
            if os.path.exists(pairings):
                pages.append(pairings)
            if os.path.exists(results):
                pages.append(results)
        # 追加：每位选手个人战绩图（放在最后），文件名形如 "01_比赛名_选手.png"
        try:
            file_list = [f for f in os.listdir(out_dir) if f.lower().endswith('.png')]
            def is_player_img(fn: str) -> bool:
                # 以两位数字开头 + 下划线，后接比赛名 + 下划线
                if len(fn) < 5 or not fn[:2].isdigit() or fn[2] != '_':
                    return False
                rest = fn[3:]
                return rest.startswith(f"{safe}_")
            player_imgs = sorted([f for f in file_list if is_player_img(f)])
            for f in player_imgs:
                pages.append(os.path.join(out_dir, f))
        except Exception:
            pass

        if not pages:
            print("[提示] 未找到可用于‘总排名与对阵情况’的页面图片。")
            return
        pdf_path = os.path.join(out_dir, f"{safe}_总排名与对阵情况.pdf")
        self.export_pdf_from_images(pages, pdf_path)
        print(f"已生成‘总排名与对阵情况’PDF：{pdf_path}")

    def ensure_out_dir(self):
        if self.out_dir is None:
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base = self._safe_filename(self.tournament_name) or '瑞士轮赛事'
            # 添加地点/组别后缀（如有）
            suffixes = []
            if getattr(self, 'tournament_location', None):
                suffixes.append(self._safe_filename(self.tournament_location))
            if getattr(self, 'tournament_category', None):
                suffixes.append(self._safe_filename(self.tournament_category))
            if suffixes:
                base = f"{base}_{'_'.join(suffixes)}"
            self.out_dir = os.path.join(os.getcwd(), f"{base}_{stamp}")
            os.makedirs(self.out_dir, exist_ok=True)
        return self.out_dir

    def _choose_bye_player(self, candidates):
        """选择轮空选手：在候选人中选最低积分，若有多名最低积分则随机选择；优先选择尚未轮空者。"""
        if not candidates:
            return None
        # 过滤掉已取消资格者
        cand = [p for p in candidates if not p.disqualified]
        if not cand:
            # 回退为原候选列表中的随机一个（避免返回 None 导致调用处错误）
            return random.choice(candidates) if candidates else None
        # 优先尚未轮空者
        no_byes = [p for p in cand if not p.bye]
        pool = no_byes if no_byes else cand
        # 找到最低分并随机选择同分者
        min_score = min(p.score for p in pool)
        mins = [p for p in pool if p.score == min_score]
        if not mins:
            # 回退保护
            return random.choice(pool)
        choice = random.choice(mins)
        return choice

    def apply_pending_disqualifications(self):
        """在显示完本轮排名后，应用待处理的取消资格（仅在仍有后续轮次时有效）。"""
        if not getattr(self, "_pending_disqualifications", None):
            return
        # 若还有后续轮次，执行取消资格；若比赛已结束，则清空即可
        if self.current_round < self.rounds:
            for name in list(self._pending_disqualifications):
                self.disqualify_player(name)
                self._pending_disqualifications.discard(name)
        else:
            self._pending_disqualifications.clear()

def main():
    try:
        tournament = SwissTournament()
        tournament.run_tournament()
    except Exception as e:
        # 运行时有未捕获异常时打印信息，便于调试
        import traceback
        print("程序运行发生错误：", e)
        traceback.print_exc()

if __name__ == "__main__":
    main()